from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (
    QApplication, QWidget, QMainWindow, QPushButton, QLabel,
    QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QLineEdit, QStackedWidget, QGridLayout, QMessageBox, QDialog
)
import sys
import os, glob
import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class Ui_CourtStats(object):
    def setupUi(self, CourtStats):
        CourtStats.setObjectName("CourtStats")
        CourtStats.resize(800, 612)
        self.centralwidget = QtWidgets.QWidget(CourtStats)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.setObjectName("verticalLayout")
        self.stackedWidget = QtWidgets.QStackedWidget(self.centralwidget)
        self.stackedWidget.setObjectName("stackedWidget")
        self.MainPage = QtWidgets.QWidget()
        self.MainPage.setObjectName("MainPage")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.MainPage)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setSpacing(0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.frame = QtWidgets.QFrame(self.MainPage)
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.frame)
        self.horizontalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_2.setSpacing(0)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.frame_2 = QtWidgets.QFrame(self.frame)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.frame_2.sizePolicy().hasHeightForWidth())
        self.frame_2.setSizePolicy(sizePolicy)
        self.frame_2.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setLineWidth(0)
        self.frame_2.setMidLineWidth(0)
        self.frame_2.setObjectName("frame_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.frame_2)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setSpacing(0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.label = QtWidgets.QLabel(self.frame_2)
        self.label.setStyleSheet("QLabel {\n"
"    font-family: \"Segoe UI\", \"Helvetica Neue\", Arial, sans-serif;\n"
"    font-size: 56px;\n"
"    font-weight: bold;\n"
"    color: black;\n"
"}")
        self.label.setObjectName("label")
        self.verticalLayout_2.addWidget(self.label, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        spacerItem = QtWidgets.QSpacerItem(20, 300, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        self.verticalLayout_2.addItem(spacerItem)
        self.newGame_main = QtWidgets.QPushButton(self.frame_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.newGame_main.sizePolicy().hasHeightForWidth())
        self.newGame_main.setSizePolicy(sizePolicy)
        self.newGame_main.setStyleSheet("QPushButton {\n"
"    font-family: \"Segoe UI\", \"Helvetica Neue\", Arial, sans-serif;\n"
"    font-size: 20px;\n"
"    font-weight: 600;\n"
"    color: #ffffff;\n"
"    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,\n"
"                                stop:0 #0072ff, stop:1 #00c6ff);\n"
"    border: none;\n"
"    border-radius: 10px;\n"
"    padding: 10px 28px;\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,\n"
"                                stop:0 #3399ff, stop:1 #33ccff);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background: qlineargradient(x1:0, y1:1, x2:1, y2:0,\n"
"                                stop:0 #0057cc, stop:1 #0099cc);\n"
"    transform: scale(0.97);\n"
"}")
        self.newGame_main.setObjectName("newGame_main")
        self.verticalLayout_2.addWidget(self.newGame_main, 0, QtCore.Qt.AlignHCenter)
        spacerItem1 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        self.verticalLayout_2.addItem(spacerItem1)
        self.loadGame_main = QtWidgets.QPushButton(self.frame_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.loadGame_main.sizePolicy().hasHeightForWidth())
        self.loadGame_main.setSizePolicy(sizePolicy)
        self.loadGame_main.setStyleSheet("QPushButton {\n"
"    font-family: \"Segoe UI\", \"Helvetica Neue\", Arial, sans-serif;\n"
"    font-size: 20px;\n"
"    font-weight: 600;\n"
"    color: #ffffff;\n"
"    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,\n"
"                                stop:0 #0072ff, stop:1 #00c6ff);\n"
"    border: none;\n"
"    border-radius: 10px;\n"
"    padding: 10px 28px;\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,\n"
"                                stop:0 #3399ff, stop:1 #33ccff);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background: qlineargradient(x1:0, y1:1, x2:1, y2:0,\n"
"                                stop:0 #0057cc, stop:1 #0099cc);\n"
"    transform: scale(0.97);\n"
"}")
        self.loadGame_main.setObjectName("loadGame_main")
        self.verticalLayout_2.addWidget(self.loadGame_main, 0, QtCore.Qt.AlignHCenter)
        spacerItem2 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        self.verticalLayout_2.addItem(spacerItem2)
        self.exit_main = QtWidgets.QPushButton(self.frame_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.exit_main.sizePolicy().hasHeightForWidth())
        self.exit_main.setSizePolicy(sizePolicy)
        self.exit_main.setStyleSheet("QPushButton {\n"
"    font-family: \"Segoe UI\", \"Helvetica Neue\", Arial, sans-serif;\n"
"    font-size: 20px;\n"
"    font-weight: 600;\n"
"    color: #ffffff;\n"
"    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,\n"
"                                stop:0 #ff1e56, stop:1 #ff7b00);\n"
"    border: none;\n"
"    border-radius: 10px;\n"
"    padding: 10px 28px;\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,\n"
"                                stop:0 #ff4d6d, stop:1 #ff9933);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background: qlineargradient(x1:0, y1:1, x2:1, y2:0,\n"
"                                stop:0 #cc0033, stop:1 #e65c00);\n"
"    transform: scale(0.97);\n"
"}")
        self.exit_main.setObjectName("exit_main")
        self.verticalLayout_2.addWidget(self.exit_main, 0, QtCore.Qt.AlignHCenter)
        spacerItem3 = QtWidgets.QSpacerItem(20, 50, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        self.verticalLayout_2.addItem(spacerItem3)
        self.horizontalLayout_2.addWidget(self.frame_2)
        self.horizontalLayout.addWidget(self.frame)
        self.stackedWidget.addWidget(self.MainPage)
        self.GameSetup = QtWidgets.QWidget()
        self.GameSetup.setObjectName("GameSetup")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.GameSetup)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.label_8 = QtWidgets.QLabel(self.GameSetup)
        font = QtGui.QFont()
        font.setFamily("Segoe UI")
        font.setPointSize(24)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_8.setFrameShadow(QtWidgets.QFrame.Raised)
        self.label_8.setLineWidth(0)
        self.label_8.setObjectName("label_8")
        self.verticalLayout_3.addWidget(self.label_8)
        self.frame_3 = QtWidgets.QFrame(self.GameSetup)
        self.frame_3.setMaximumSize(QtCore.QSize(16777215, 50))
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.horizontalLayout_10 = QtWidgets.QHBoxLayout(self.frame_3)
        self.horizontalLayout_10.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_10.setSpacing(0)
        self.horizontalLayout_10.setObjectName("horizontalLayout_10")
        self.frame_13 = QtWidgets.QFrame(self.frame_3)
        self.frame_13.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_13.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_13.setObjectName("frame_13")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout(self.frame_13)
        self.horizontalLayout_3.setContentsMargins(0, 0, -1, 0)
        self.horizontalLayout_3.setSpacing(0)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_2 = QtWidgets.QLabel(self.frame_13)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_3.addWidget(self.label_2)
        self.leagueName_input = QtWidgets.QLineEdit(self.frame_13)
        self.leagueName_input.setObjectName("leagueName_input")
        self.horizontalLayout_3.addWidget(self.leagueName_input)
        self.horizontalLayout_10.addWidget(self.frame_13)
        self.frame_14 = QtWidgets.QFrame(self.frame_3)
        self.frame_14.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_14.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_14.setObjectName("frame_14")
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout(self.frame_14)
        self.horizontalLayout_9.setContentsMargins(0, 0, -1, 0)
        self.horizontalLayout_9.setSpacing(0)
        self.horizontalLayout_9.setObjectName("horizontalLayout_9")
        spacerItem4 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_9.addItem(spacerItem4)
        self.label_5 = QtWidgets.QLabel(self.frame_14)
        self.label_5.setObjectName("label_5")
        self.horizontalLayout_9.addWidget(self.label_5)
        self.gameNo_input = QtWidgets.QLineEdit(self.frame_14)
        self.gameNo_input.setObjectName("gameNo_input")
        self.horizontalLayout_9.addWidget(self.gameNo_input, 0, QtCore.Qt.AlignLeft)
        self.horizontalLayout_10.addWidget(self.frame_14)
        self.verticalLayout_3.addWidget(self.frame_3)
        self.frame_4 = QtWidgets.QFrame(self.GameSetup)
        self.frame_4.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout(self.frame_4)
        self.horizontalLayout_4.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_4.setSpacing(0)
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.frame_5 = QtWidgets.QFrame(self.frame_4)
        self.frame_5.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_5.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_5.setObjectName("frame_5")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout(self.frame_5)
        self.verticalLayout_4.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_4.setSpacing(0)
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.frame_7 = QtWidgets.QFrame(self.frame_5)
        self.frame_7.setMaximumSize(QtCore.QSize(16777215, 50))
        self.frame_7.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_7.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_7.setObjectName("frame_7")
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout(self.frame_7)
        self.horizontalLayout_6.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_6.setSpacing(0)
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.label_3 = QtWidgets.QLabel(self.frame_7)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_6.addWidget(self.label_3)
        self.homeTeam_input = QtWidgets.QLineEdit(self.frame_7)
        self.homeTeam_input.setObjectName("homeTeam_input")
        self.horizontalLayout_6.addWidget(self.homeTeam_input)
        self.verticalLayout_4.addWidget(self.frame_7)
        self.frame_8 = QtWidgets.QFrame(self.frame_5)
        self.frame_8.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_8.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_8.setObjectName("frame_8")
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout(self.frame_8)
        self.horizontalLayout_7.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_7.setSpacing(0)
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        self.homeTeam_table = QtWidgets.QTableWidget(self.frame_8)
        self.homeTeam_table.setObjectName("homeTeam_table")
        self.homeTeam_table.setColumnCount(4)
        self.homeTeam_table.setRowCount(0)

        # Home team table
        self.homeTeam_table.setHorizontalHeaderLabels(["ID", "Player Name", "Jersey", "Status"])

        self.horizontalLayout_7.addWidget(self.homeTeam_table)
        self.verticalLayout_4.addWidget(self.frame_8)
        self.horizontalLayout_13 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_13.setObjectName("horizontalLayout_13")
        self.addPlayer_home = QtWidgets.QPushButton(self.frame_5)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.addPlayer_home.sizePolicy().hasHeightForWidth())
        self.addPlayer_home.setSizePolicy(sizePolicy)
        self.addPlayer_home.setObjectName("addPlayer_home")
        self.horizontalLayout_13.addWidget(self.addPlayer_home)
        self.PushButton = QtWidgets.QPushButton(self.frame_5)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.PushButton.sizePolicy().hasHeightForWidth())
        self.PushButton.setSizePolicy(sizePolicy)
        self.PushButton.setObjectName("PushButton")
        self.horizontalLayout_13.addWidget(self.PushButton)
        self.verticalLayout_4.addLayout(self.horizontalLayout_13)
        self.horizontalLayout_4.addWidget(self.frame_5)
        self.frame_6 = QtWidgets.QFrame(self.frame_4)
        self.frame_6.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_6.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_6.setObjectName("frame_6")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.frame_6)
        self.verticalLayout_5.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_5.setSpacing(0)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.frame_10 = QtWidgets.QFrame(self.frame_6)
        self.frame_10.setMaximumSize(QtCore.QSize(16777215, 50))
        self.frame_10.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_10.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_10.setObjectName("frame_10")
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout(self.frame_10)
        self.horizontalLayout_5.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_5.setSpacing(0)
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.label_4 = QtWidgets.QLabel(self.frame_10)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_5.addWidget(self.label_4)
        self.awayTeam_input = QtWidgets.QLineEdit(self.frame_10)
        self.awayTeam_input.setObjectName("awayTeam_input")
        self.horizontalLayout_5.addWidget(self.awayTeam_input)
        self.verticalLayout_5.addWidget(self.frame_10)
        self.frame_11 = QtWidgets.QFrame(self.frame_6)
        self.frame_11.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_11.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_11.setObjectName("frame_11")
        self.horizontalLayout_8 = QtWidgets.QHBoxLayout(self.frame_11)
        self.horizontalLayout_8.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_8.setSpacing(0)
        self.horizontalLayout_8.setObjectName("horizontalLayout_8")
        self.awayTeam_table = QtWidgets.QTableWidget(self.frame_11)
        self.awayTeam_table.setObjectName("awayTeam_table")
        self.awayTeam_table.setColumnCount(4)
        self.awayTeam_table.setRowCount(0)

        # Away Team table
        self.awayTeam_table.setHorizontalHeaderLabels(["ID", "Player Name", "Jersey", "Status"])

        self.horizontalLayout_8.addWidget(self.awayTeam_table)
        self.verticalLayout_5.addWidget(self.frame_11)
        self.horizontalLayout_14 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_14.setObjectName("horizontalLayout_14")
        self.addPlayer_away = QtWidgets.QPushButton(self.frame_6)
        self.addPlayer_away.setObjectName("addPlayer_away")
        self.horizontalLayout_14.addWidget(self.addPlayer_away)
        self.removePlayer_away = QtWidgets.QPushButton(self.frame_6)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.removePlayer_away.sizePolicy().hasHeightForWidth())
        self.removePlayer_away.setSizePolicy(sizePolicy)
        self.removePlayer_away.setObjectName("removePlayer_away")
        self.horizontalLayout_14.addWidget(self.removePlayer_away)
        self.verticalLayout_5.addLayout(self.horizontalLayout_14)
        self.horizontalLayout_4.addWidget(self.frame_6)
        self.verticalLayout_3.addWidget(self.frame_4)
        self.horizontalLayout_15 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_15.setObjectName("horizontalLayout_15")
        self.gameSetup_back = QtWidgets.QPushButton(self.GameSetup)
        self.gameSetup_back.setObjectName("gameSetup_back")
        self.horizontalLayout_15.addWidget(self.gameSetup_back)
        self.startGameStat = QtWidgets.QPushButton(self.GameSetup)
        self.startGameStat.setObjectName("startGameStat")
        self.horizontalLayout_15.addWidget(self.startGameStat)
        self.verticalLayout_3.addLayout(self.horizontalLayout_15)
        self.stackedWidget.addWidget(self.GameSetup)
        self.Stats_page = QtWidgets.QWidget()
        self.Stats_page.setObjectName("Stats_page")
        self.verticalLayout_8 = QtWidgets.QVBoxLayout(self.Stats_page)
        self.verticalLayout_8.setObjectName("verticalLayout_8")
        self.frame_15 = QtWidgets.QFrame(self.Stats_page)
        self.frame_15.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_15.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_15.setObjectName("frame_15")
        self.verticalLayout_9 = QtWidgets.QVBoxLayout(self.frame_15)
        self.verticalLayout_9.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_9.setSpacing(0)
        self.verticalLayout_9.setObjectName("verticalLayout_9")
        self.label_6 = QtWidgets.QLabel(self.frame_15)
        font = QtGui.QFont()
        font.setFamily("Segoe UI")
        font.setPointSize(24)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.verticalLayout_9.addWidget(self.label_6, 0, QtCore.Qt.AlignHCenter)
        spacerItem5 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        self.verticalLayout_9.addItem(spacerItem5)
        self.leagueName = QtWidgets.QLabel(self.frame_15)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.leagueName.setFont(font)
        self.leagueName.setObjectName("leagueName")
        self.verticalLayout_9.addWidget(self.leagueName, 0, QtCore.Qt.AlignHCenter)
        spacerItem6 = QtWidgets.QSpacerItem(20, 5, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        self.verticalLayout_9.addItem(spacerItem6)
        self.gameNo = QtWidgets.QLabel(self.frame_15)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.gameNo.setFont(font)
        self.gameNo.setScaledContents(True)
        self.gameNo.setObjectName("gameNo")
        self.verticalLayout_9.addWidget(self.gameNo, 0, QtCore.Qt.AlignHCenter)
        self.verticalLayout_8.addWidget(self.frame_15)
        self.frame_17 = QtWidgets.QFrame(self.Stats_page)
        self.frame_17.setMinimumSize(QtCore.QSize(0, 250))
        self.frame_17.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_17.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_17.setObjectName("frame_17")
        self.horizontalLayout_11 = QtWidgets.QHBoxLayout(self.frame_17)
        self.horizontalLayout_11.setObjectName("horizontalLayout_11")
        self.frame_22 = QtWidgets.QFrame(self.frame_17)
        self.frame_22.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_22.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_22.setObjectName("frame_22")
        self.verticalLayout_11 = QtWidgets.QVBoxLayout(self.frame_22)
        self.verticalLayout_11.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_11.setSpacing(0)
        self.verticalLayout_11.setObjectName("verticalLayout_11")
        self.homeTeam = QtWidgets.QLabel(self.frame_22)
        self.homeTeam.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.homeTeam.setFont(font)
        self.homeTeam.setObjectName("homeTeam")
        self.verticalLayout_11.addWidget(self.homeTeam, 0, QtCore.Qt.AlignHCenter)
        self.frame_16 = QtWidgets.QFrame(self.frame_22)
        self.frame_16.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_16.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_16.setObjectName("frame_16")
        self.gridLayout = QtWidgets.QGridLayout(self.frame_16)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setSpacing(0)
        self.gridLayout.setObjectName("gridLayout")
        self.HT6 = QtWidgets.QPushButton(self.frame_16)
        self.HT6.setObjectName("HT6")
        self.gridLayout.addWidget(self.HT6, 1, 5, 1, 1)
        self.HT10 = QtWidgets.QPushButton(self.frame_16)
        self.HT10.setObjectName("HT10")
        self.gridLayout.addWidget(self.HT10, 3, 1, 1, 1)
        self.HT2 = QtWidgets.QPushButton(self.frame_16)
        self.HT2.setObjectName("HT2")
        self.gridLayout.addWidget(self.HT2, 0, 3, 1, 1)
        self.HT1 = QtWidgets.QPushButton(self.frame_16)
        self.HT1.setObjectName("HT1")
        self.gridLayout.addWidget(self.HT1, 0, 1, 1, 1)
        self.HT4 = QtWidgets.QPushButton(self.frame_16)
        self.HT4.setObjectName("HT4")
        self.gridLayout.addWidget(self.HT4, 1, 1, 1, 1)
        self.HT3 = QtWidgets.QPushButton(self.frame_16)
        self.HT3.setObjectName("HT3")
        self.gridLayout.addWidget(self.HT3, 0, 5, 1, 1)
        self.HT12 = QtWidgets.QPushButton(self.frame_16)
        self.HT12.setObjectName("HT12")
        self.gridLayout.addWidget(self.HT12, 3, 5, 1, 1)
        self.HT9 = QtWidgets.QPushButton(self.frame_16)
        self.HT9.setObjectName("HT9")
        self.gridLayout.addWidget(self.HT9, 2, 5, 1, 1)
        spacerItem7 = QtWidgets.QSpacerItem(5, 0, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem7, 0, 2, 1, 1)
        self.HT7 = QtWidgets.QPushButton(self.frame_16)
        self.HT7.setObjectName("HT7")
        self.gridLayout.addWidget(self.HT7, 2, 1, 1, 1)
        self.HT5 = QtWidgets.QPushButton(self.frame_16)
        self.HT5.setObjectName("HT5")
        self.gridLayout.addWidget(self.HT5, 1, 3, 1, 1)
        self.HT8 = QtWidgets.QPushButton(self.frame_16)
        self.HT8.setObjectName("HT8")
        self.gridLayout.addWidget(self.HT8, 2, 3, 1, 1)
        self.HT11 = QtWidgets.QPushButton(self.frame_16)
        self.HT11.setObjectName("HT11")
        self.gridLayout.addWidget(self.HT11, 3, 3, 1, 1)
        spacerItem8 = QtWidgets.QSpacerItem(5, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem8, 0, 4, 1, 1)
        self.verticalLayout_11.addWidget(self.frame_16)
        self.horizontalLayout_11.addWidget(self.frame_22)
        spacerItem9 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_11.addItem(spacerItem9)
        self.frame_21 = QtWidgets.QFrame(self.frame_17)
        self.frame_21.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_21.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_21.setObjectName("frame_21")
        self.verticalLayout_10 = QtWidgets.QVBoxLayout(self.frame_21)
        self.verticalLayout_10.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_10.setSpacing(0)
        self.verticalLayout_10.setObjectName("verticalLayout_10")
        self.awayTeam = QtWidgets.QLabel(self.frame_21)
        self.awayTeam.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.awayTeam.setFont(font)
        self.awayTeam.setObjectName("awayTeam")
        self.verticalLayout_10.addWidget(self.awayTeam, 0, QtCore.Qt.AlignHCenter)
        self.frame_20 = QtWidgets.QFrame(self.frame_21)
        self.frame_20.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_20.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_20.setObjectName("frame_20")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.frame_20)
        self.gridLayout_2.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_2.setSpacing(0)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.AT2 = QtWidgets.QPushButton(self.frame_20)
        self.AT2.setObjectName("AT2")
        self.gridLayout_2.addWidget(self.AT2, 0, 3, 1, 1)
        self.AT5 = QtWidgets.QPushButton(self.frame_20)
        self.AT5.setObjectName("AT5")
        self.gridLayout_2.addWidget(self.AT5, 1, 3, 1, 1)
        self.AT11 = QtWidgets.QPushButton(self.frame_20)
        self.AT11.setObjectName("AT11")
        self.gridLayout_2.addWidget(self.AT11, 3, 3, 1, 1)
        self.AT3 = QtWidgets.QPushButton(self.frame_20)
        self.AT3.setObjectName("AT3")
        self.gridLayout_2.addWidget(self.AT3, 0, 5, 1, 1)
        self.AT1 = QtWidgets.QPushButton(self.frame_20)
        self.AT1.setObjectName("AT1")
        self.gridLayout_2.addWidget(self.AT1, 0, 1, 1, 1)
        self.AT9 = QtWidgets.QPushButton(self.frame_20)
        self.AT9.setObjectName("AT9")
        self.gridLayout_2.addWidget(self.AT9, 2, 5, 1, 1)
        self.AT8 = QtWidgets.QPushButton(self.frame_20)
        self.AT8.setObjectName("AT8")
        self.gridLayout_2.addWidget(self.AT8, 2, 3, 1, 1)
        self.AT10 = QtWidgets.QPushButton(self.frame_20)
        self.AT10.setObjectName("AT10")
        self.gridLayout_2.addWidget(self.AT10, 3, 1, 1, 1)
        self.AT12 = QtWidgets.QPushButton(self.frame_20)
        self.AT12.setObjectName("AT12")
        self.gridLayout_2.addWidget(self.AT12, 3, 5, 1, 1)
        spacerItem10 = QtWidgets.QSpacerItem(5, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem10, 0, 2, 1, 1)
        self.AT7 = QtWidgets.QPushButton(self.frame_20)
        self.AT7.setObjectName("AT7")
        self.gridLayout_2.addWidget(self.AT7, 2, 1, 1, 1)
        self.AT6 = QtWidgets.QPushButton(self.frame_20)
        self.AT6.setObjectName("AT6")
        self.gridLayout_2.addWidget(self.AT6, 1, 5, 1, 1)
        self.AT4 = QtWidgets.QPushButton(self.frame_20)
        self.AT4.setObjectName("AT4")
        self.gridLayout_2.addWidget(self.AT4, 1, 1, 1, 1)
        spacerItem11 = QtWidgets.QSpacerItem(5, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem11, 0, 4, 1, 1)
        self.verticalLayout_10.addWidget(self.frame_20)
        self.horizontalLayout_11.addWidget(self.frame_21)
        self.verticalLayout_8.addWidget(self.frame_17)
        self.frame_18 = QtWidgets.QFrame(self.Stats_page)
        self.frame_18.setMinimumSize(QtCore.QSize(600, 0))
        self.frame_18.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_18.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_18.setObjectName("frame_18")
        self.verticalLayout_12 = QtWidgets.QVBoxLayout(self.frame_18)
        self.verticalLayout_12.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_12.setSpacing(0)
        self.verticalLayout_12.setObjectName("verticalLayout_12")
        self.playByplay_log = QtWidgets.QListWidget(self.frame_18)
        self.playByplay_log.setObjectName("playByplay_log")
        self.verticalLayout_12.addWidget(self.playByplay_log)
        self.verticalLayout_8.addWidget(self.frame_18, 0, QtCore.Qt.AlignHCenter)
        self.frame_19 = QtWidgets.QFrame(self.Stats_page)
        self.frame_19.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_19.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_19.setObjectName("frame_19")
        self.horizontalLayout_12 = QtWidgets.QHBoxLayout(self.frame_19)
        self.horizontalLayout_12.setObjectName("horizontalLayout_12")
        self.statsPage_back = QtWidgets.QPushButton(self.frame_19)
        self.statsPage_back.setObjectName("statsPage_back")
        self.horizontalLayout_12.addWidget(self.statsPage_back)
        self.statsPage_undo = QtWidgets.QPushButton(self.frame_19)
        self.statsPage_undo.setObjectName("statsPage_undo")
        self.horizontalLayout_12.addWidget(self.statsPage_undo)
        self.statsPage_view = QtWidgets.QPushButton(self.frame_19)
        self.statsPage_view.setObjectName("statsPage_view")
        self.horizontalLayout_12.addWidget(self.statsPage_view)
        self.verticalLayout_8.addWidget(self.frame_19)
        self.stackedWidget.addWidget(self.Stats_page)
        self.verticalLayout.addWidget(self.stackedWidget)
        CourtStats.setCentralWidget(self.centralwidget)

        self.retranslateUi(CourtStats)
        self.stackedWidget.setCurrentIndex(0)

        self.newGame_main.clicked.connect(self.go_to_newGame)
        self.startGameStat.clicked.connect(self.start_newGame)
        self.gameSetup_back.clicked.connect(self.back_to_home)
        self.statsPage_back.clicked.connect(self.back_to_previous)
        self.addPlayer_away.clicked.connect(lambda: self.add_row(self.awayTeam_table))
        self.addPlayer_home.clicked.connect(lambda: self.add_row(self.homeTeam_table))
        self.removePlayer_away.clicked.connect(lambda: self.remove_row(self.awayTeam_table))
        self.PushButton.clicked.connect(lambda: self.remove_row(self.homeTeam_table))
        self.startGameStat.clicked.connect(self.start_newGame)

        QtCore.QMetaObject.connectSlotsByName(CourtStats)

    def go_to_newGame (self):
        self.stackedWidget.setCurrentIndex(1)

        # Create main directory
        base_dir = os.path.join(os.getcwd(), "StatsRecord")
        os.makedirs(base_dir, exist_ok=True)

        # Create timestamped subfolder
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"StatsRecord_{timestamp}.xlsx"
        file_path = os.path.join(base_dir, file_name)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "GameStats"

        # Header info
        ws["A1"] = "League:"
        ws["A2"] = "Game No:"
        ws["A3"] = "Created:"
        ws["B3"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Column headers
        headers = [
                "PlayerID", "Player_names", "Jersey_no", "Team", "Score",
                "Assists", "Rebounds", "Steals", "Blocks", "FTM", "FTA", "3PM", "3PA"
        ]

        for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True)

        # Set reasonable column widths
        for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 15

        # Save workbook
        wb.save(file_path)

        print(f"Excel file created: {file_path}")

    def start_newGame (self):
        self.stackedWidget.setCurrentIndex(2)
        self.rename_label()
        self.save_to_excel()

        file_path = get_latest_stats_file()
        if not file_path:
                print("No stats file found.")
                return

        home_team, away_team = get_team_names(file_path)
        if not home_team or not away_team:
                print("Team names missing in Excel.")
                return

        home_jerseys, away_jerseys = load_team_jerseys(file_path, home_team, away_team)
        self.update_team_buttons(home_jerseys, away_jerseys)
    
#     def update_team_buttons(self, home_jerseys, away_jerseys):
#         # Update Home Team buttons
#         for i in range(1, len(home_jerseys) + 1):
#                 try:
#                         ht_btn = getattr(self, f"HT{i}")  # remove .ui
#                         ht_btn.setText(str(home_jerseys[i - 1]))
#                         ht_btn.setEnabled(True)
#                 except AttributeError:
#                         print(f"Home button HT{i} not found in UI.")

#         # Update Away Team buttons
#         for i in range(1, len(away_jerseys) + 1):
#                 try:
#                         at_btn = getattr(self, f"AT{i}")  # remove .ui
#                         at_btn.setText(str(away_jerseys[i - 1]))
#                         at_btn.setEnabled(True)
#                 except AttributeError:
#                         print(f"Away button AT{i} not found in UI.")

#     def update_team_buttons(self, home_jerseys, away_jerseys):
#         # Swap the assignment
#         # Home jerseys go to HT buttons
#         for i in range(1, len(home_jerseys) + 1):
#                 try:
#                         at_btn = getattr(self, f"AT{i}")  # was HT before
#                         at_btn.setText(str(home_jerseys[i - 1]))
#                         at_btn.setEnabled(True)
#                 except AttributeError:
#                         print(f"Away button AT{i} not found in UI.")

#         # Away jerseys go to AT buttons
#         for i in range(1, len(away_jerseys) + 1):
#                 try:
#                         ht_btn = getattr(self, f"HT{i}")  # was AT before
#                         ht_btn.setText(str(away_jerseys[i - 1]))
#                         ht_btn.setEnabled(True)
#                 except AttributeError:
#                         print(f"Home button HT{i} not found in UI.")

    def update_team_buttons(self, home_jerseys, away_jerseys):
        max_buttons = 12  # assuming you have HT1-12 and AT1-12

        # Home Team buttons
        for i in range(1, max_buttons + 1):
                try:
                        ht_btn = getattr(self, f"AT{i}")
                        if i <= len(home_jerseys):
                                ht_btn.setText(str(home_jerseys[i - 1]))
                                ht_btn.setEnabled(True)     # enable used buttons
                                ht_btn.setVisible(True)     # show used buttons
                        else:
                                ht_btn.setText("")          # clear text
                                ht_btn.setEnabled(False)    # disable unused
                                ht_btn.setVisible(False)    # hide unused
                except AttributeError:
                        print(f"Home button HT{i} not found in UI.")

        # Away Team buttons
        for i in range(1, max_buttons + 1):
                try:
                        at_btn = getattr(self, f"HT{i}")
                        if i <= len(away_jerseys):
                                at_btn.setText(str(away_jerseys[i - 1]))
                                at_btn.setEnabled(True)
                                at_btn.setVisible(True)
                        else:
                                at_btn.setText("")
                                at_btn.setEnabled(False)
                                at_btn.setVisible(False)
                except AttributeError:
                        print(f"Away button AT{i} not found in UI.")

    def rename_label (self):
        leagueName_txt = self.leagueName_input.text()
        gameNo_txt = self.gameNo_input.text()
        homeTeam_txt = self.homeTeam_input.text()
        awayTeam_txt = self.awayTeam_input.text()

        self.leagueName.setText(leagueName_txt)
        self.gameNo.setText(gameNo_txt)
        self.homeTeam.setText(homeTeam_txt)
        self.awayTeam.setText(awayTeam_txt)

    def save_to_excel(self):
        folder_path = os.path.join(os.getcwd(), "StatsRecord")

        # Ensure folder exists
        if not os.path.exists(folder_path):
                print("StatsRecord folder not found!")
                return

        # Get all Excel files in the folder
        excel_files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
        if not excel_files:
                print("No Excel files found in StatsRecord folder!")
                return

        # Pick the latest file based on modification time
        latest_file = max(excel_files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)))
        file_path = os.path.join(folder_path, latest_file)
        print(f"Loading Excel file: {file_path}")

        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active  # Assuming your sheet is the first one

        # Update metadata
        ws["B1"] = f"{self.leagueName.text()}"
        ws["B2"] = f"{self.gameNo.text()}"
        ws["B3"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Start writing player data from row 5
        current_row = 5

        # Home team
        home_team = self.homeTeam_input.text()
        for row in range(self.homeTeam_table.rowCount()):
                ws.cell(row=current_row, column=1, value=self.homeTeam_table.item(row, 0).text())
                ws.cell(row=current_row, column=2, value=self.homeTeam_table.item(row, 1).text())
                ws.cell(row=current_row, column=3, value=self.homeTeam_table.item(row, 2).text())
                ws.cell(row=current_row, column=4, value=home_team)
                # Initialize stats columns to 0
                for col in range(5, 14):
                        ws.cell(row=current_row, column=col, value=0)
                current_row += 1

        # Away team
        away_team = self.awayTeam_input.text()
        for row in range(self.awayTeam_table.rowCount()):
                ws.cell(row=current_row, column=1, value=self.awayTeam_table.item(row, 0).text())
                ws.cell(row=current_row, column=2, value=self.awayTeam_table.item(row, 1).text())
                ws.cell(row=current_row, column=3, value=self.awayTeam_table.item(row, 2).text())
                ws.cell(row=current_row, column=4, value=away_team)
                # Initialize stats columns to 0
                for col in range(5, 14):
                        ws.cell(row=current_row, column=col, value=0)
                current_row += 1

        # Save workbook
        # wb.save(file_path)
        # print(f"Data populated successfully in: {file_path}")

        try:
                wb.save(file_path)
        except PermissionError:
                print(f"Cannot save Excel file '{file_path}'. It may be open in Excel.")

    def back_to_previous (self):
        self.stackedWidget.setCurrentIndex(1)

    def back_to_home (self):
        self.stackedWidget.setCurrentIndex(0)

    def add_row(self, table):
        row = table.rowCount()
        table.insertRow(row)
        table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
        table.setItem(row, 1, QTableWidgetItem(""))
        table.setItem(row, 2, QTableWidgetItem(""))

#     def remove_row(self, table):
#         selected = table.currentRow()
#         if selected >= 0:
#                 table.removeRow(selected)

    def remove_row(self, table):
        selected = table.currentRow()
        if selected >= 0:
                # Remove row from QTable
                table.removeRow(selected)

                folder_path = os.path.join(os.getcwd(), "StatsRecord")

                # Pick the latest file
                excel_files = [f for f in os.listdir(folder_path)
                        if f.endswith(".xlsx") and not f.startswith("~$")]
                if not excel_files:
                        print("No Excel files found!")
                        return

                latest_file = max(excel_files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)))
                file_path = os.path.join(folder_path, latest_file)

                wb = load_workbook(file_path)
                ws = wb.active

                # Adjust row for Excel (assuming data starts at row 5)
                excel_row = selected + 5
                ws.delete_rows(excel_row)

                # wb.save(file_path)
                # print(f"Row {selected} removed from both QTable and Excel")

                try:
                        wb.save(file_path)
                        print(f"Row {selected} removed from both QTable and Excel")
                except PermissionError:
                        print(f"Permission denied: Could not update Excel. Please close '{latest_file}' in Excel and try again.")


    def retranslateUi(self, CourtStats):
        _translate = QtCore.QCoreApplication.translate
        CourtStats.setWindowTitle(_translate("CourtStats", "CourtStats"))
        self.label.setText(_translate("CourtStats", "CourtStats"))
        self.newGame_main.setText(_translate("CourtStats", "New Game"))
        self.loadGame_main.setText(_translate("CourtStats", "Load Game"))
        self.exit_main.setText(_translate("CourtStats", "Exit"))
        self.label_8.setText(_translate("CourtStats", "Team Setup"))
        self.label_2.setText(_translate("CourtStats", "League Name: "))
        self.label_5.setText(_translate("CourtStats", "Game No.: "))
        self.label_3.setText(_translate("CourtStats", "Home Team: "))
        self.addPlayer_home.setText(_translate("CourtStats", "(+) Add player"))
        self.PushButton.setText(_translate("CourtStats", "(-) Remove player"))
        self.label_4.setText(_translate("CourtStats", "Away Team: "))
        self.addPlayer_away.setText(_translate("CourtStats", "(+) Add player"))
        self.removePlayer_away.setText(_translate("CourtStats", "(-) Remove player"))
        self.gameSetup_back.setText(_translate("CourtStats", "Back"))
        self.startGameStat.setText(_translate("CourtStats", "Start"))
        self.label_6.setText(_translate("CourtStats", "Game Stats"))
        self.leagueName.setText(_translate("CourtStats", "LeagueName"))
        self.gameNo.setText(_translate("CourtStats", "GameNo"))
        self.homeTeam.setText(_translate("CourtStats", "HomeTeam"))
        self.HT6.setText(_translate("CourtStats", "HT6"))
        self.HT10.setText(_translate("CourtStats", "HT10"))
        self.HT2.setText(_translate("CourtStats", "HT2"))
        self.HT1.setText(_translate("CourtStats", "HT1"))
        self.HT4.setText(_translate("CourtStats", "HT4"))
        self.HT3.setText(_translate("CourtStats", "HT3"))
        self.HT12.setText(_translate("CourtStats", "HT12"))
        self.HT9.setText(_translate("CourtStats", "HT9"))
        self.HT7.setText(_translate("CourtStats", "HT7"))
        self.HT5.setText(_translate("CourtStats", "HT5"))
        self.HT8.setText(_translate("CourtStats", "HT8"))
        self.HT11.setText(_translate("CourtStats", "HT11"))
        self.awayTeam.setText(_translate("CourtStats", "AwayTeam"))
        self.AT2.setText(_translate("CourtStats", "AT2"))
        self.AT5.setText(_translate("CourtStats", "AT5"))
        self.AT11.setText(_translate("CourtStats", "AT11"))
        self.AT3.setText(_translate("CourtStats", "AT3"))
        self.AT1.setText(_translate("CourtStats", "AT1"))
        self.AT9.setText(_translate("CourtStats", "AT9"))
        self.AT8.setText(_translate("CourtStats", "AT8"))
        self.AT10.setText(_translate("CourtStats", "AT10"))
        self.AT12.setText(_translate("CourtStats", "AT12"))
        self.AT7.setText(_translate("CourtStats", "AT7"))
        self.AT6.setText(_translate("CourtStats", "AT6"))
        self.AT4.setText(_translate("CourtStats", "AT4"))
        self.statsPage_back.setText(_translate("CourtStats", "Back"))
        self.statsPage_undo.setText(_translate("CourtStats", "Undo"))
        self.statsPage_view.setText(_translate("CourtStats", "View"))


def load_team_jerseys(file_path, home_team_name, away_team_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    home_jerseys, away_jerseys = [], []

    team_col = None
    jersey_col = None
    header_row = None

    # scan first 10 rows to find the header
    for r in range(1, 11):
        for i, cell in enumerate(ws[r], start=1):
            if not cell.value:
                continue
            value = str(cell.value).strip().lower()
            if value == "team":
                team_col = i
                header_row = r
            elif value in ["jersey_no", "jersey", "jersey no"]:
                jersey_col = i
                header_row = r
        if team_col and jersey_col:
            break

    if team_col is None or jersey_col is None:
        raise ValueError(
            f"Missing 'Team' or 'Jersey_no' columns. "
            f"Checked first 10 rows. Detected headers: {[cell.value for row in ws.iter_rows(min_row=1, max_row=10) for cell in row]}"
        )

#     for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
#         team = row[team_col - 1]
#         jersey = row[jersey_col - 1]
#         if not team or jersey is None:
#             continue
#         if team == home_team_name:
#             home_jerseys.append(str(jersey))
#         elif team == away_team_name:
#             away_jerseys.append(str(jersey))

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        team = str(row[team_col - 1]).strip().lower()
        jersey = row[jersey_col - 1]
        if not team or jersey is None:
                continue
        if team == home_team_name.strip().lower():
                home_jerseys.append(str(jersey))
        elif team == away_team_name.strip().lower():
                away_jerseys.append(str(jersey))

    print("Home jerseys:", home_jerseys)
    print("Away jerseys:", away_jerseys)
    return home_jerseys, away_jerseys
    
    
def get_latest_stats_file():
        folder = os.path.join(os.getcwd(), "StatsRecord")
        files = glob.glob(os.path.join(folder, "StatsRecord_*.xlsx"))
        if not files:
                return None
        return max(files, key=os.path.getctime)

def get_team_names(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    header_row = 4  # adjust if needed
    team_col = None

    # Find the 'Team' column
    for i, cell in enumerate(ws[header_row], start=1):
        if cell.value == "Team":
            team_col = i
            break  # only break after found

    if team_col is None:
        raise ValueError("Could not find 'Team' column in header row")

    teams = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        team = row[team_col - 1]
        if team:
            teams.add(team)

    teams = list(teams)
    if not teams:
        raise ValueError("No team names found in file")

    if len(teams) >= 2:
        return teams[0], teams[1]
    else:
        return teams[0], None
